from datetime import datetime
import glob
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pandas as pd
import shutil
import streamlit as st
import sys
import trimesh
from dotenv import load_dotenv
import tkinter as tk
from tkinter import ttk
import numpy as np
from PIL import Image as PILImage, ImageDraw, ImageFont
import io


load_dotenv()
tempCarpet = os.getenv('carpetaTemporal')
scaledFolder=os.getenv('scaledFolder')
savingFolder=f"{tempCarpet}\{scaledFolder}"

def getActualGfolder(year,month):
    pathGoogle=os.getenv('GooglePath')
    folder=os.path.join(pathGoogle,year,month)
    if not os.path.exists(folder):
        os.mkdir(folder)
    return folder

def getNextQuotaNumber(projectFolder):
    maxNumber=0
    for file in os.listdir(projectFolder):
        d = os.path.join(projectFolder, file)
        if os.path.isdir(d):
            x = d.split('_')
            if len(x) >= 3 and x[2].isdigit():
                number = int(x[2])
                if number >= maxNumber:
                    maxNumber = number
    maxNumber=maxNumber+1
    print(maxNumber)
    return maxNumber

def cleanOldFiles():
    # Clearing files, if exist
    if not os.path.exists(tempCarpet):
        os.mkdir(tempCarpet)

    print("DeletingFiles")
    # Delete files in tempDir
    file_list = glob.glob(os.path.join(tempCarpet, '*'))
    for file_path in file_list:
        try:
            if os.path.isfile(file_path):
                print(f"Deleting file: {file_path}")
                os.remove(file_path)
            elif os.path.isdir(file_path):
                print(f"Deleting directory: {file_path}")
                shutil.rmtree(file_path)
            else:
                print(f"Not a file or directory: {file_path}")
        except Exception as e:
            print(f"Error deleting file {file_path}: {e}")

    files = os.listdir()
    for file in files:
        if file.endswith(".zip"):
            os.remove(file)

def areImagesGenerated():
    for path in os.listdir(tempCarpet + "\\"):
        # check if current path is a file
        if os.path.isfile(os.path.join(tempCarpet, path)):
            nombreA = os.path.join(tempCarpet, path)
            if (nombreA.endswith(".jpg")):
                return True
    return False

def combine_images(images):
    """
    Combina una lista de imágenes en una sola imagen.
    """
    # Obtener el tamaño de la imagen combinada
    max_width = max(img.width for img in images)
    total_height = sum(img.height for img in images)

    # Crear una nueva imagen combinada con el tamaño adecuado
    combined_image = PILImage.new("RGB", (max_width, total_height), color="white")
    y_offset = 0

    # Superponer cada imagen en la imagen combinada
    for img in images:
        combined_image.paste(img, (0, y_offset))
        y_offset += img.height

    return combined_image

def areQuotaGenerated():
    for path in os.listdir(tempCarpet + "\\"):
        # check if current path is a file
        if os.path.isfile(os.path.join(tempCarpet, path)):
            nombreA = os.path.join(tempCarpet, path)
            if (nombreA.endswith(".xlsx")):
                return True
    return False

def cleanActualProjectCarpet(path):
    file_list = glob.glob(os.path.join(path, '*'))
    for file_path in file_list:
        os.remove(file_path)
        
def toggle_image(image_bytes, caption, key):
    show_image = st.checkbox(f"Mostrar {caption}", key=key)
    if show_image:
        st.image(image_bytes, caption=caption)
        
def cargar_datos_distribuidor(marca):
    # Leer el archivo Excel
    df = pd.read_excel('resources\InventarioDistribuidores.xlsx')

    # Filtrar los datos por la marca seleccionada
    datos_marca = df[df['Marca'] == marca]

    # Crear una lista para almacenar los datos de la marca seleccionada
    datos_seleccionados = []

    # Iterar sobre los datos de la marca seleccionada y almacenar la descripción y el precio de distribuidor
    for index, row in datos_marca.iterrows():
        descripcion = row['Descripción']
        precio_distribuidor = row['Precio Distribuidor']
        venta_sugerida = row['Venta Sugerida']

        # Agregar los datos a la lista
        datos_seleccionados.append({
            'Descripción': descripcion,
            'Precio Distribuidor': precio_distribuidor,
            'Venta Sugerida': venta_sugerida
        })

    return datos_seleccionados
        
def MainApp():
    fecha_actual = datetime.now()
    current_year = str(fecha_actual.year)
    current_month = str(fecha_actual.month)
    if 'Estado' not in st.session_state:
        st.session_state['Estado'] = 0
    if 'quotaNumber' not in st.session_state:
        st.session_state['quotaNumber'] = 1

    rootdir = getActualGfolder(current_year, current_month)
    # Obtener el siguiente número de cotización
    if st.session_state['Estado'] == 0:
        cleanOldFiles()
        st.session_state['quotaNumber'] = getNextQuotaNumber(rootdir)
        st.session_state['Estado'] = 1
        st.rerun() 
    else:
        st.title("Cotizador Captis")
        st.markdown('---')
        quotaNumber = st.session_state['quotaNumber']
        name2Quota = f"C{current_year}_{current_month}_{quotaNumber}"
        st.markdown(f"Folio de la cotizacion: {name2Quota}")
        if st.session_state['Estado'] == 2:
            st.write("Se ha subido correctamente, puedes generar otra cotizacion")
            if st.button("Generar una nueva cotización"):
                st.session_state['Estado'] = 0
                st.rerun() 
            return
        elif not st.session_state['Estado'] == 1:
            st.write("Error")
            return
        else:
            # Datos generales
            cliente = st.text_input('Nombre del cliente')
            description = st.text_input('Descripcion corta')
            empresa = st.text_input('Empresa')
            moneda = st.selectbox('Moneda:', ['MXN', 'USD'])
            concepto = st.selectbox('Tipo de servicio:', ['Servicio de manufactura aditiva', 'Diseño o modelado 3D', 'Venta de insumos', 'Capacitación'])
            estudiante = st.checkbox("Es estudiante?")
            materialesInfo = pd.read_csv("resources\Materiales.csv")
            
            # Mostrar campos específicos según el tipo de servicio seleccionado
            if concepto == 'Servicio de manufactura aditiva':
                envio = st.radio(
                    'Selecciona el tipo de envio',
                    options=['Envio nacional Paquetexpress', 'Envio a punto de venta', 'Envio local Monterrey'])

                adressShipping = os.getenv('adressShipping')
                shippingToSalesPoint = os.getenv('shippingToSalesPoint')
                expressShipping = os.getenv('expressShipping')
                
                shipping_cost = 0
                if envio == 'Envio nacional Paquetexpress':
                    shipping_cost = adressShipping  # Puedes ajustar este valor según tus necesidades
                elif envio == 'Envio a punto de venta':
                    shipping_cost = shippingToSalesPoint  # Puedes ajustar este valor según tus necesidades
                else:
                    shipping_cost = expressShipping  # Puedes ajustar este valor según tus necesidades
                
                shipping_cost = float(shipping_cost)
                incluirElEnvio = st.checkbox("Incluir el costo del envio en el precio(Beta)", False)
                
                opciones = materialesInfo['Material']
                material = st.radio(
                    "Elige el material👇",
                    opciones,
                    key="visibility",
                )
                relleno = 10

                if "Resina" in material:
                    relleno = int(st.slider('Cantidad de relleno',
                                    min_value=25, max_value=100, format="%i"))
                else:
                    relleno = int(st.slider('Cantidad de relleno',
                                        min_value=10, max_value=100, format="%i"))

                try2Fix = st.checkbox('Se repararán los archivos')

                args = [material, relleno, try2Fix]
                formats=[]
                
                for i in trimesh.available_formats():
                    formats.append(f".{i}")
                    # Loading data
                
                data = st.file_uploader(f"Inserta los archivos para la cotización, solo {formats}", type=formats, accept_multiple_files=True)
                st.markdown('---')

                # Resto del código para analizar, mostrar imágenes, preparar cotización, etc.
                if len(data) > 0:
                    # Converting files if requested
                    if st.button('Analiza los archivos'):
                        cleanOldFiles()
                        # Escritura de archivos en tempCarpet
                        for cad in data:
                            source = os.path.join(tempCarpet, cad.name)
                            dest = os.path.splitext(source)[0] + '.stl'
                            with open(source, "wb") as f:
                                f.write(cad.getbuffer())
                            
                            st.write(f'Cargando: {os.path.basename(source)}...')
                        
                        st.write('Analizando, ...')
                        
                        if incluirElEnvio:
                            command2Send = f'{sys.executable} CotizadorCMD.py {args[0]} {args[1]} {args[2]} {int(shipping_cost)*100}'
                            print(command2Send)
                            os.system(command2Send)
                        else:
                            command2Send = f'{sys.executable} CotizadorCMD.py {args[0]} {args[1]} {args[2]} {0}'
                            print(command2Send)
                            os.system(command2Send)
                        
                        dest = os.path.join(tempCarpet, 'info.csv')
                        dest2 = os.path.join(tempCarpet, 'easy.csv')
                        if os.path.exists(dest) and os.path.exists(dest2):
                            st.markdown('Se han analizado los archivos correctamente')
                            df = pd.read_csv(dest)
                            easy2read = pd.read_csv(dest2)
                            st.dataframe(easy2read)
                            st.dataframe(df)
                        else:
                            st.markdown('Ha habido un error')

                        # Procesamiento de imágenes generadas
                        if areImagesGenerated():
                            if st.button('Mostrar imagenes de archivos'):
                                all_image_paths = []
                                processed_image_paths = set()  # Conjunto para mantener un registro de las imágenes procesadas
                                for path in os.listdir(tempCarpet):
                                    # Verificar si la ruta actual es un archivo de imagen
                                    if os.path.isfile(os.path.join(tempCarpet, path)):
                                        nombreA = os.path.join(tempCarpet, path)
                                        if (nombreA.endswith(".jpg")):
                                            # Verificar si la imagen ya ha sido procesada
                                            if nombreA not in processed_image_paths:
                                                img = PILImage.open(nombreA)
                                                all_image_paths.append((nombreA, img))
                                                processed_image_paths.add(nombreA)

                                # Combinar todas las imágenes en una sola
                                combined_image = combine_images([img for _, img in all_image_paths])

                                # Guardar la imagen combinada
                                join_image_path = os.path.join(tempCarpet, "joined_image.jpg")
                                combined_image.save(join_image_path)

                                # Mostrar la imagen combinada
                                st.image(combined_image, caption="joined_image.jpg")

                                num_images = len(all_image_paths)
                                st.write(f"Total de imágenes encontradas: {num_images}")

                                for img_path, _ in all_image_paths:
                                    st.image(PILImage.open(img_path), caption=os.path.basename(img_path), width=200)
                            
                            # Eliminar la imagen combinada de la lista de imágenes para que no se repita
                                processed_image_paths.discard(join_image_path)
                                
                        dest = os.path.join(tempCarpet, 'easy.csv')
                        if os.path.exists(dest):
                            if st.button('Preparar la cotizacion'):
                                st.markdown(f'Generando...')
                                infoCsv = os.path.join(tempCarpet, "easy.csv")
                                df = pd.read_csv(infoCsv)
                                path2Template = 'resources\Template.xlsx'
                                outputFilename = os.path.join(tempCarpet, name2Quota + '.xlsx')

                                templateWorkbook = load_workbook(path2Template)
                                destination_worksheet = templateWorkbook['Quota']
                                totals_worksheet=templateWorkbook['Datos']
                                if len(cliente)==0:
                                    cliente="Cliente "+name2Quota
                                destination_worksheet['B3'].value = cliente
                                destination_worksheet['B4'].value = concepto
                                destination_worksheet['G1'].value = name2Quota
                                destination_worksheet['G2'].value = fecha_actual.strftime("%d / %m / %y")
                                destination_worksheet['G4'].value = moneda
                                destination_worksheet['G4'].value = empresa


                                df = df.assign(Cantidad=1)
                                df = df.assign(Subtotal=df['CostoConIVA'])

                                start_row = 12
                                for row in dataframe_to_rows(df, index=False, header=False):
                                    for idx, value in enumerate(row, start=1):
                                        destination_worksheet.cell(row=start_row, column=idx, value=value)
                                    start_row += 1

                                getMaterial=df['Material'][0]
                                print("Informacion del material selecionado:")
                                print(getMaterial)
                                dfmat = pd.read_csv("resources\Materiales.csv")
                                infoMaterial = dfmat[dfmat["Material"] == getMaterial]
                                print(infoMaterial)
                                
                                if incluirElEnvio:
                                    shipping_cost=0
                                destination_worksheet.cell(row=start_row, column=1).value=envio
                                shipping_cost = shipping_cost  
                                destination_worksheet.cell(row=start_row, column=4).value = shipping_cost
                                destination_worksheet.cell(row=start_row, column=5).value = shipping_cost * 1.16
                                destination_worksheet.cell(row=start_row, column=6).value = 1
                                destination_worksheet.cell(row=start_row, column=7).value = shipping_cost * 1.16
                                start_row += 1

                                total = df['CostoConIVA'].sum()+float(shipping_cost)*1.16

                                if estudiante:
                                    destination_worksheet.cell(row=start_row, column=1).value="Descuento del 15%"
                                    destination_worksheet.cell(row=start_row, column=7).value=-total*0.15
                                    
                                infoCsv = os.path.join(tempCarpet, "info.csv")
                                dftotals = pd.read_csv(infoCsv)
                                timeInDays = f"{np.round((dftotals['TiempoEstimado'].sum()/60/24+1)*1.61,0)}"
                                totals_worksheet.cell(row=3, column=3).value=timeInDays

                                st.markdown(f"Total de tiempo: {timeInDays} dias habiles")
                                templateWorkbook.save(outputFilename)
                                templateWorkbook.close()

                        #Append Images 
                        image_path = f"{tempCarpet}/joined_image.jpg"
                        print(image_path)
                        img=Image(image_path)
                        destination_worksheet.add_image(img, 'A5')

                        #Compresing Files
                        shutil.make_archive(name2Quota, 'zip',tempCarpet)
                        st.markdown(f'Generado a las {datetime.now()}')
                        zipName=f'{name2Quota}.zip'
                        if os.path.exists(zipName):
                            with open(zipName, 'rb') as f:
                                st.download_button('Descargar', f, file_name=zipName)

                            if(areQuotaGenerated()):
                                if st.button("Subir a la carpeta de proyectos"):
                                    #Movinig to a project carpet to be generated
                                    projectFolder=os.path.join(rootdir,name2Quota)
                                    if not os.path.exists(projectFolder):
                                        os.mkdir(projectFolder)
                                    else:
                                        cleanActualProjectCarpet(projectFolder)

                                    for file_name in os.listdir(tempCarpet):
                                        # construct full file path
                                        source = os.path.join(tempCarpet,file_name)
                                        destination = os.path.join(projectFolder,file_name)
                                        # copy only files
                                        if os.path.isfile(source):
                                            shutil.copy(source, destination)
                                            print('Copiado', file_name)
                                    st.session_state['Estado'] = 2
                                    st.success("Uploaded")
                                    if(st.button("Confirmar")):
                                        st.rerun()
                        else:
                            st.markdown('Aun no se han analizado los archivos')
                    
            
            elif concepto == 'Diseño o modelado 3D':
                # Muestra un campo de carga de archivos para que los usuarios carguen una imagen
                uploaded_files = st.file_uploader("Cargar imagen", type=["jpg", "png", "jpeg"], accept_multiple_files=True)
                # Comprueba si se ha cargado un archivo
                if uploaded_files is not None:
                    # Itera sobre cada archivo cargado
                    for i, uploaded_file in enumerate(uploaded_files):
                        # Lee los datos del archivo cargado como bytes
                        image_bytes = uploaded_file.read()
                        # Muestra el botón de alternancia para la previsualización de la imagen
                        toggle_image(image_bytes, f"Imagen {i+1}", f"checkbox_{i+1}")
                   
    
            elif concepto == 'Venta de insumos':
                st.subheader("Insumo")

                # Selección de marca
                marca_seleccionada = st.selectbox("Selecciona la marca:", ["COLOR PLUS", "SUNLU", "Otro"])

                if marca_seleccionada == "Otro":
                    if 'num_insumos' not in st.session_state:
                        st.session_state['num_insumos'] = 1

                    st.subheader("Insumo")

                    # Si el número de insumos es menor a 1, establecerlo a 1
                    if st.session_state['num_insumos'] < 1:
                        st.session_state['num_insumos'] = 1

                    insumos_data = []

                    for i in range(st.session_state['num_insumos']):
                        col1, col2, col3, col4 = st.columns([1, 1, 1, 1])

                        with col1:
                            descripciondeproducto = st.text_input(f'Descripción del producto {i+1}:', "")
                                
                        with col2:
                            cantidad = st.text_input(f'Cantidad {i+1}:', "")

                        with col3:
                            costo = st.text_input(f"Costo (cuánto nos cuesta) {i+1}:", "")

                            # Validación del formato de la cantidad monetaria
                            if costo:
                                try:
                                    costo = float(costo.replace(",", ""))  # Elimina las comas si las hay
                                    if costo < 0:
                                        st.error(f"El costo del producto {i+1} debe ser un valor positivo.")
                                except ValueError:
                                    st.error(f"Formato de costo del producto {i+1} incorrecto. Introduce un número válido.")

                        with col4:
                            # Botón para eliminar la fila
                            if st.session_state['num_insumos'] > 1:
                                if st.button(f"Eliminar", help="Eliminar producto", key=f"delete_button_{i}"):
                                    # Eliminar el insumo correspondiente al índice actual
                                    st.session_state['num_insumos'] -= 1
                                    # Acción de eliminar el producto aquí

                        # Almacenar los datos del insumo actual en la lista
                        insumos_data.append({
                            'Descripción': descripciondeproducto,
                            'Cantidad': cantidad,
                            'Costo': costo
                        })

                    # Botón para agregar otro elemento
                    if st.button("Agregar otro insumo"):
                        st.session_state['num_insumos'] += 1

                    # Botón para analizar los costos
                    if st.button("Analizar costos"):
                        # Acción de analizar costos y calcular precio al público aquí
                        st.write("Se analizarán los costos y se calculará el precio al público aquí.")
                        # Aquí deberías incluir el cálculo del precio al público y cualquier otra lógica necesaria
                        
                        # Creamos un DataFrame con los datos de todos los insumos
                        insumos_df = pd.DataFrame(insumos_data)

                        # Se calcula el precio al público agregando el 35% al costo
                        insumos_df['Precio al público'] = insumos_df['Costo'] * 1.35

                        # Se muestra el DataFrame con el cálculo del precio al público
                        st.dataframe(insumos_df)
                    
                else:
                    # Cargar y mostrar datos del distribuidor seleccionado
                    datos_distribuidor = cargar_datos_distribuidor(marca_seleccionada)

                    # Mostrar los datos según la cantidad
                    cantidad_insumos = []
                    for producto in datos_distribuidor:
                        col1, col2 = st.columns([2, 1])  # Ajustar la proporción de las columnas

                        with col1:
                            descripcion = producto["Descripción"]
                            st.write(descripcion)  # Mostrar descripción en la primera columna

                        with col2:
                            cantidad = st.number_input(f'{descripcion} cantidad', min_value=0, value=0, step=1, label_visibility='collapsed')
                            if cantidad != 0:
                                 # Calcular precios con IVA
                                precio_distribuidor_iva = producto['Precio Distribuidor'] * 1.16
                                venta_sugerida_iva = producto['Venta Sugerida'] * 1.16
                                    
                                cantidad_insumos.append({
                                    'Descripción': descripcion,
                                    'Cantidad': cantidad,
                                    'Precio Distribuidor (sin IVA)': producto['Precio Distribuidor'],
                                    'Precio Distribuidor (con IVA)': precio_distribuidor_iva,
                                    'Venta sugerida (sin IVA)': producto['Venta Sugerida'],
                                    'Venta sugerida (con IVA)': venta_sugerida_iva
                                })

                    # Si hay productos con cantidad diferente de 0, mostrar el DataFrame
                    if cantidad_insumos:
                        insumos_df = pd.DataFrame(cantidad_insumos)
                        st.write("Productos con cantidad diferente de 0:")
                        st.dataframe(insumos_df[['Descripción', 'Cantidad', 'Precio Distribuidor (sin IVA)', 'Precio Distribuidor (con IVA)', 'Venta sugerida (sin IVA)', 'Venta sugerida (con IVA)']])
                    else:
                        st.write("No hay productos con cantidad diferente de 0.")

                    # Botón para analizar los costos
                    if st.button("Preparar cotización"):
                        st.write("Se analizarán los costos y se calculará el precio al público aquí.")
                        # Aquí deberías incluir el cálculo del precio al público y cualquier otra lógica necesaria

            
            elif concepto == 'Capacitación':
                # Mostrar campos específicos para Capacitación
                # Aquí puedes agregar los campos necesarios para este tipo de servicio
                ...
         
MainApp()