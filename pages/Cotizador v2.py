import glob
import io
import os
import shutil
import sys
from datetime import datetime
from tkinter import ttk
import tkinter as tk

import numpy as np
import pandas as pd
import trimesh
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image as PILImage, ImageDraw, ImageFont
import streamlit as st

load_dotenv()
tempCarpet = os.getenv('carpetaTemporal')
scaledFolder=os.getenv('scaledFolder')
savingFolder=f"{tempCarpet}\{scaledFolder}"

def getActualGfolder(year,month):
    pathGoogle=os.getenv('GooglePath')
    folder=os.path.join(pathGoogle,year,month)
    if not os.path.exists(folder):
        os.mkdir(folder)
    return folder

def getNextQuotaNumber(projectFolder):
    maxNumber = 0
    for file in os.listdir(projectFolder):
        d = os.path.join(projectFolder, file)
        if os.path.isdir(d):
            parts = file.split('_')
            if len(parts) >= 3 and parts[2].isdigit():
                number = int(parts[2])
                if number > maxNumber:
                    maxNumber = number
    nextNumber = maxNumber + 1
    print(nextNumber)
    return nextNumber

def cleanOldFiles(tempCarpet):
    # Ensure the temporary directory exists
    os.makedirs(tempCarpet, exist_ok=True)
    
    print("Deleting files in temporary directory")
    # Delete files and directories in tempCarpet
    for file_path in glob.glob(os.path.join(tempCarpet, '*')):
        try:
            if os.path.isfile(file_path):
                print(f"Deleting file: {file_path}")
                os.remove(file_path)
            elif os.path.isdir(file_path):
                print(f"Deleting directory: {file_path}")
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Error deleting {file_path}: {e}")

    # Delete all .zip files in the current directory
    for file in os.listdir():
        if file.endswith(".zip"):
            try:
                print(f"Deleting zip file: {file}")
                os.remove(file)
            except Exception as e:
                print(f"Error deleting zip file {file}: {e}")

def areImagesGenerated(tempCarpet):
    for path in os.listdir(tempCarpet):
        full_path = os.path.join(tempCarpet, path)
        if os.path.isfile(full_path) and full_path.endswith(".jpg"):
            return True
    return False

def combine_images(images):
    """
    Combines a list of images into a single image.

    Args:
        images (list of PIL.Image): List of images to combine.

    Returns:
        PIL.Image: The combined image.
    """
    if not images:
        return None  # or raise an exception if an empty list is not expected

    # Determine the size of the combined image
    max_width = max(img.width for img in images)
    total_height = sum(img.height for img in images)

    # Create a new combined image with the appropriate size
    combined_image = PILImage.new("RGB", (max_width, total_height), color="white")
    y_offset = 0

    # Overlay each image onto the combined image
    for img in images:
        combined_image.paste(img, (0, y_offset))
        y_offset += img.height

    return combined_image

def areQuotaGenerated():
    for path in os.listdir(tempCarpet + "\\"):
        # check if current path is a file
        if os.path.isfile(os.path.join(tempCarpet, path)):
            nombreA = os.path.join(tempCarpet, path)
            if (nombreA.endswith(".xlsx")):
                return True
    return False

def cleanActualProjectCarpet(path):
    file_list = glob.glob(os.path.join(path, '*'))
    for file_path in file_list:
        os.remove(file_path)
        
def toggle_image(image_bytes, caption, key):
    show_image = st.checkbox(f"Mostrar {caption}", key=key)
    if show_image:
        st.image(image_bytes, caption=caption)

def MainApp():
    fecha_actual = datetime.now()
    current_year = str(fecha_actual.year)
    current_month = str(fecha_actual.month)
    if 'Estado' not in st.session_state:
        st.session_state['Estado'] = 0
    if 'quotaNumber' not in st.session_state:
        st.session_state['quotaNumber'] = 1

    rootdir = getActualGfolder(current_year, current_month)
    # Obtener el siguiente número de cotización
    if st.session_state['Estado'] == 0:
        cleanOldFiles(tempCarpet=tempCarpet)
        st.session_state['quotaNumber'] = getNextQuotaNumber(rootdir)
        print(st.session_state['quotaNumber'])
        st.session_state['Estado'] = 1
        st.rerun() 
    else:
        st.title("Cotizador Captis")
        st.markdown('---')
        quotaNumber = st.session_state['quotaNumber']
        name2Quota = f"C{current_year}_{current_month}_{quotaNumber}"
        st.markdown(f"Folio de la cotizacion: {name2Quota}")
        if st.session_state['Estado'] == 2:
            st.write("Se ha subido correctamente, puedes generar otra cotizacion")
            if st.button("Generar una nueva cotización"):
                st.session_state['Estado'] = 0
                st.rerun() 
            return
        elif not st.session_state['Estado'] == 1:
            st.write("Error")
            return
        else:
            # Datos generales
            cliente = st.text_input('Nombre del cliente')
            empresa = st.text_input('Empresa')
            estudiante = st.checkbox("Es estudiante?")
            materialesInfo = pd.read_csv("resources\Materiales.csv")
            
            envio = st.radio(
                'Selecciona el tipo de envio',
                options=['Envio nacional Paquetexpress', 'Envio a punto de venta', 'Envio local Monterrey'])

            adressShipping = os.getenv('adressShipping')
            shippingToSalesPoint = os.getenv('shippingToSalesPoint')
            expressShipping = os.getenv('expressShipping')
            
            shipping_cost = 0
            if envio == 'Envio nacional Paquetexpress':
                shipping_cost = adressShipping  # Puedes ajustar este valor según tus necesidades
            elif envio == 'Envio a punto de venta':
                shipping_cost = shippingToSalesPoint  # Puedes ajustar este valor según tus necesidades
            else:
                shipping_cost = expressShipping  # Puedes ajustar este valor según tus necesidades
            
            shipping_cost = float(shipping_cost)
            incluirElEnvio = st.checkbox("Incluir el costo del envio en el precio(Beta)", False)
            
            opciones = materialesInfo['Material']
            material = st.radio(
                "Elige el material👇",
                opciones,
                key="visibility",
            )
            relleno = 10

            if "Resina" in material:
                relleno = int(st.slider('Cantidad de relleno',
                                min_value=25, max_value=100, format="%i"))
            else:
                relleno = int(st.slider('Cantidad de relleno',
                                    min_value=10, max_value=100, format="%i"))

            try2Fix = st.checkbox('Se repararán los archivos')

            args = [material, relleno, try2Fix]
            formats=[]
            
            for i in trimesh.available_formats():
                formats.append(f".{i}")
                # Loading data
            
            data = st.file_uploader(f"Inserta los archivos para la cotización, solo {formats}", type=formats, accept_multiple_files=True)
            st.markdown('---')

            # Resto del código para analizar, mostrar imágenes, preparar cotización, etc.
            if len(data) > 0:
                # Converting files if requested
                if st.button('Analiza los archivos'):
                    cleanOldFiles(tempCarpet=tempCarpet)
                    # Escritura de archivos en tempCarpet
                    for cad in data:
                        source = os.path.join(tempCarpet, cad.name)
                        dest = os.path.splitext(source)[0] + '.stl'
                        with open(source, "wb") as f:
                            f.write(cad.getbuffer())
                        
                        st.write(f'Cargando: {os.path.basename(source)}...')
                    
                    st.write('Analizando, ...')
                    
                    if incluirElEnvio:
                        command2Send = f'{sys.executable} CotizadorCMD.py {args[0]} {args[1]} {args[2]} {int(shipping_cost)*100}'
                        print(command2Send)
                        os.system(command2Send)
                    else:
                        command2Send = f'{sys.executable} CotizadorCMD.py {args[0]} {args[1]} {args[2]} {0}'
                        print(command2Send)
                        os.system(command2Send)
                    
                    dest = os.path.join(tempCarpet, 'info.csv')
                    dest2 = os.path.join(tempCarpet, 'easy.csv')
                    if os.path.exists(dest) and os.path.exists(dest2):
                        st.markdown('Se han analizado los archivos correctamente')
                        df = pd.read_csv(dest)
                        easy2read = pd.read_csv(dest2)
                        st.dataframe(easy2read)
                        st.dataframe(df)
                        # Procesamiento de imágenes generadas
                        if areImagesGenerated(tempCarpet=tempCarpet):
                            all_image_paths = []
                            processed_image_paths = set()  # Conjunto para mantener un registro de las imágenes procesadas
                            for path in os.listdir(tempCarpet):
                                # Verificar si la ruta actual es un archivo de imagen
                                if os.path.isfile(os.path.join(tempCarpet, path)):
                                    nombreA = os.path.join(tempCarpet, path)
                                    if (nombreA.endswith(".jpg")):
                                        # Verificar si la imagen ya ha sido procesada
                                        if nombreA not in processed_image_paths:
                                            img = PILImage.open(nombreA)
                                            all_image_paths.append((nombreA, img))
                                            processed_image_paths.add(nombreA)


                            num_images = len(all_image_paths)
                            st.write(f"Total de imágenes encontradas: {num_images}")

                            for img_path, _ in all_image_paths:
                                st.image(PILImage.open(img_path), caption=os.path.basename(img_path), width=200)
                    else:
                        st.markdown('Ha habido un error')


                dest = os.path.join(tempCarpet, 'easy.csv')
                if os.path.exists(dest):
                    if st.button('Preparar la cotizacion'):
                        st.markdown(f'Generando...')
                        infoCsv = os.path.join(tempCarpet, "easy.csv")
                        df = pd.read_csv(infoCsv)
                        path2Template = 'resources\Template.xlsx'
                        outputFilename = os.path.join(tempCarpet, name2Quota + '.xlsx')

                        templateWorkbook = load_workbook(path2Template)
                        destination_worksheet = templateWorkbook['Quota']
                        totals_worksheet=templateWorkbook['Datos']
                        if len(cliente)==0:
                            cliente="Cliente "+name2Quota
                        destination_worksheet['B3'].value = cliente
                        destination_worksheet['B4'].value = "Servicio de manufactura aditiva"
                        destination_worksheet['G1'].value = name2Quota
                        destination_worksheet['G2'].value = fecha_actual.strftime("%d / %m / %y")
                        destination_worksheet['G4'].value = "MXN"
                        destination_worksheet['G4'].value = empresa

                        df = df.assign(Cantidad=1)
                        df = df.assign(Subtotal=df['CostoConIVA'])

                        start_row = 12
                        for row in dataframe_to_rows(df, index=False, header=False):
                            for idx, value in enumerate(row, start=1):
                                destination_worksheet.cell(row=start_row, column=idx, value=value)
                            start_row += 1

                        getMaterial=df['Material'][0]
                        print("Informacion del material selecionado:")
                        print(getMaterial)
                        dfmat = pd.read_csv("resources\Materiales.csv")
                        infoMaterial = dfmat[dfmat["Material"] == getMaterial]
                        print(infoMaterial)
                        
                        if incluirElEnvio:
                            shipping_cost=0
                        destination_worksheet.cell(row=start_row, column=1).value=envio
                        shipping_cost = shipping_cost  
                        destination_worksheet.cell(row=start_row, column=4).value = shipping_cost
                        destination_worksheet.cell(row=start_row, column=5).value = shipping_cost * 1.16
                        destination_worksheet.cell(row=start_row, column=6).value = 1
                        destination_worksheet.cell(row=start_row, column=7).value = shipping_cost * 1.16
                        start_row += 1

                        total = df['CostoConIVA'].sum()+float(shipping_cost)*1.16

                        if estudiante:
                            destination_worksheet.cell(row=start_row, column=1).value="Descuento del 15%"
                            destination_worksheet.cell(row=start_row, column=7).value=-total*0.15
                            
                        infoCsv = os.path.join(tempCarpet, "info.csv")
                        dftotals = pd.read_csv(infoCsv)
                        timeInDays = f"{np.round((dftotals['TiempoEstimado'].sum()/60/24+1)*1.61,0)}"
                        totals_worksheet.cell(row=3, column=3).value=timeInDays
                        st.markdown(f"Total de tiempo: {timeInDays} dias habiles")

                        
                        #Append Images 
                        image_path = f"{tempCarpet}/joined_image.jpg"
                        print(image_path)
                        img=Image(image_path)
                        destination_worksheet.add_image(img, 'A5')

                        templateWorkbook.save(outputFilename)
                        templateWorkbook.close()

                        #Compresing Files
                        shutil.make_archive(name2Quota, 'zip',tempCarpet)
                        st.markdown(f'Generado a las {datetime.now()}')
                            
                        zipName=f'{name2Quota}.zip'
                        if os.path.exists(zipName):
                            with open(zipName, 'rb') as f:
                                st.download_button('Descargar', f, file_name=zipName)
                        else:
                            st.markdown('Aun no se han analizado los archivos')
                            
                    if(areQuotaGenerated()):
                        if st.button("Subir a la carpeta de proyectos"):
                            #Movinig to a project carpet to be generated
                            projectFolder=os.path.join(rootdir,name2Quota)
                            if not os.path.exists(projectFolder):
                                os.mkdir(projectFolder)
                            else:
                                cleanActualProjectCarpet(projectFolder)

                            for file_name in os.listdir(tempCarpet):
                                # construct full file path
                                source = os.path.join(tempCarpet,file_name)
                                destination = os.path.join(projectFolder,file_name)
                                # copy only files
                                if os.path.isfile(source):
                                    shutil.copy(source, destination)
                                    print('Copiado', file_name)
                            st.session_state['Estado'] = 2
                            st.success("Uploaded")
                            if(st.button("Confirmar")):
                                st.rerun()

                            
MainApp()